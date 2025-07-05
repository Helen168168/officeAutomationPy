"""
每周日自动创建周计划
使用pandas复制数据，用openpyxl处理合并单元格
"""

import os
import cn2an
import schedule
import time
from copy import copy
from openpyxl import load_workbook
from datetime import datetime, timedelta

def autoEnterData(source_path, target_path):
    # 检查源文件是否存在
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"源文件不存在: {source_path}")

        # 获取当前日期和时间
        now = datetime.now()
        print(f"{now} 开始执行每周自动导入任务...")
    wb_target = load_workbook(target_path)
    # 获取目标sheet处理格式
    newSheetName = getNewSheetName(source_path)
    # 创建目标Sheet
    target_sheet = wb_target.create_sheet(newSheetName)

    # 用openpyxl读取源文件的信息
    wb_source = load_workbook(source_path)

    # 获取所有的sheet名称
    sheetNames = getAllSheets(source_path)
    lastSheetName = sheetNames[len(sheetNames) - 1]
    sheetNameObj = wb_source[lastSheetName]


    # 复制合并单元格
    for merge in sheetNameObj.merged_cells.ranges:
        target_sheet.merge_cells(str(merge))

    #复制单元格样式和内容
    for row in sheetNameObj.iter_rows():
        for cell in row:
            if cell.value is not None and '周计划' in cell.value:
                cell.value = getNextDateRange() + '周计划'
            newCell = target_sheet.cell(
                row = cell.row,
                column = cell.column,
                value=cell.value
            )

            if cell.style:
                newCell.font = copy(cell.font)
                newCell.border = copy(cell.border)
                newCell.fill = copy(cell.fill)
                newCell.number_format = cell.number_format
                newCell.protection = copy(cell.protection)
                newCell.alignment = copy(cell.alignment)

        # 复制列宽和行高
        for col_idx, col_dim in sheetNameObj.column_dimensions.items():
            target_sheet.column_dimensions[col_idx].width = col_dim.width
            print(f'cell_width = {col_dim.width}')

        for row_idx, row_dim in sheetNameObj.row_dimensions.items():
            target_sheet.row_dimensions[row_idx].height = row_dim.height if row_dim.height is not None else 80
            print(f'cell_height = {target_sheet.row_dimensions[row_idx].height}')

    # 4. 保存目标文件
    wb_target.save(target_path)
    print(f"恭喜！已经自动录入周计划数据，距离成功又进了一步！")

# 定义文件路径（使用原始字符串避免转义问题）
source_path = r"E:\资料库\2025周计划集合.xlsx"  # 源文件路径
target_path = source_path

# 将数字转换为大写
def tansformNumberToLetter(number):
    upLetter = cn2an.an2cn(number)
    return upLetter

# 获取Excel文件中所有Sheet名称
def getAllSheets(file_path):
    # 用openpyxl读取源文件的信息
    wb_source = load_workbook(file_path)
    # 获取所有的sheet名称
    sheetNames = [sheetName for sheetName in wb_source.sheetnames]
    return sheetNames

# 获取需要新增Sheet的名称
def getNewSheetName(source_path):
    # 读取源文件中所有Sheet
    sheet_names = getAllSheets(source_path)
    lens = len(sheet_names)
    upLetter = tansformNumberToLetter(lens + 1)
    sheet_name = f"第{upLetter}周"
    return sheet_name

def is_sunday():
    return datetime.now().weekday() == 6

def getNextDateRange():
    today = datetime.now().date()
    # 计算下周一（当前日期的下周一）
    next_monday = today + timedelta(days=(7 - today.weekday()))

    # 计算下周日（下周一 + 6天）
    next_sunday = next_monday + timedelta(days=6)

    next_monday = next_monday.strftime('%Y/%m/%d')
    next_sunday = next_sunday.strftime('%Y/%m/%d')

    return f'{next_monday}-{next_sunday}'

# 每周日晚上九点自动执行
def autoExecute():
    schedule.every().sunday.at("21:00").do(autoEnterData)
    while True:
        if(is_sunday() and datetime.now().hour == 21 and datetime.now().minute < 2):
            autoEnterData(source_path, target_path)
            time.sleep(60)  # 避免一小时内重复执行
        schedule.run_pending()
        time.sleep(60) # 每分钟检查一次

# autoExecute()

autoEnterData(source_path, target_path)