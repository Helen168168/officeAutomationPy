'''
 获取Excel数据
'''
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta, time

class ExcelOperation:
    def __init__(self, path=""):
        self.path = path
        self.allSheets = []  # 所有sheet对象
        self.allSheetNames = []  # 所有sheet名称
        self.sheetNameData = {}  # sheet名称和数据的映射
        pass

    def is_file_path_valid(self):
        # 检查文件是否存在
        if not os.path.exists(self.path):
            raise FileNotFoundError(f"文件不存在: {self.path}")
            return False
        return True

    def getAllSheets(self):
        if not self.is_file_path_valid():
            return
        # 读取Excel文件的信息
        wb_target = load_workbook(self.path)
        # 获取所有的sheet名称
        self.allSheetNames = [sheetName for sheetName in wb_target.sheetnames]
        # 获取所有sheet对象
        self.allSheets = [wb_target[sheetName] for sheetName in self.allSheetNames]


    # 获取所有sheet的周计划信息
    def getAllWeekPlanInfo(self):
        allCellContent = {}
        # 获取所有sheet的数据
        for sheet in self.allSheets:
            dictKey = ''
            weekName = ''
            for col in sheet.iter_cols():
                for cell in col:
                    # 去掉值为None的单元格
                    if cell.value is not None:
                        if cell.value.find('周计划') != -1 and cell.coordinate == 'A1':
                            weekName = cell.value
                            allCellContent[weekName] = {}
                            continue
                        if cell.value in ['周一', '周二', '周三', '周四', '周五', '周六', '周日']:
                            dictKey = cell.value
                            allCellContent[weekName][dictKey] = []
                            continue
                        if dictKey != '':
                            allCellContent[weekName][dictKey].append(cell.value)
        return allCellContent

    # 日期切割
    def splitDate(self, str):
        if str != '' and str != None:
            date = str.split('-')[0]
            return date

    '''
        将数据组装为对象集合的格式,便于在飞书中添加日程
        [{startTime: '', endTime: '', content: ''}]
    '''
    def splitTimePlan(self, date, plan):
        timePlan = []
        if plan != '' and plan != None:
            timeInfo = plan.split('\n\n')
            if len(timeInfo) > 0:
                for time in timeInfo:
                    periodInfo = {}
                    info = time.split(' ')
                    for index, content in enumerate(info):
                        if index == 0:
                            hourMinus = content.split('-')
                            for index, val in enumerate(hourMinus):
                                strDate = date.strftime('%Y/%m/%d') + ' ' + val
                                parsedDateTime = datetime.strptime(strDate, '%Y/%m/%d %H:%M')
                                if index == 0:
                                    periodInfo['startTime'] = parsedDateTime
                                else:
                                    periodInfo['endTime'] = parsedDateTime
                        else:
                            periodInfo['content'] = content
                    timePlan.append(periodInfo)
            return timePlan
        else:
            return []

    '''
       获取所有的日程计划和日程执行的反馈信息 
    '''
    def splitScheduleInfo(self):
        scheduleInfo = [] # 记录所以工作表的日程信息
        feedbackInfo = [] # 记录所有工作表的日程执行的反馈信息
        allCellContent = self.getAllWeekPlanInfo()
        for key, val in dict(allCellContent).items():
            startWeekDate = self.splitDate(key)
            for index, (weekName, weekPlan) in enumerate(val.items()):
                weekDate = datetime.strptime(startWeekDate, '%Y/%m/%d').date() + timedelta(days=index)
                if len(weekPlan) > 0:
                    for index, plan in enumerate(weekPlan):
                        if index == 0:
                            dayInfo = self.splitTimePlan(weekDate, plan)
                            scheduleInfo.append(dayInfo)
                        else:
                            feedbackInfo.append(plan)
        return (scheduleInfo, feedbackInfo)