from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_excel_report(data, output_path):
    """生成Excel格式报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据分析报告"

    # 添加标题
    ws['A1'] = "数据分析报告"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')

    # 写入数据
    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)

    # 设置样式
    for col in ws.columns:
        max_length = 0
        if hasattr(col[0], 'column_letter'):
            column = col[0].column_letter
        else:
            column = col[1].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal='center')
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    print(f"Excel报告已生成: { output_path }")